import requests
import requests.exceptions
from lxml import html
from bs4 import BeautifulSoup
from openpyxl import Workbook
import logging
import time
import argparse

logging.basicConfig(format='%(filename)s LINE:%(lineno)d | '
                           '[%(asctime)s]# | %(levelname)s '
                           '| %(message)s',
                    level=logging.INFO,
                    filename=u'log.log')
logger = logging.getLogger(__name__)


def get_courses_content():
    logging.info(u'start getting XML')
    url = 'https://www.coursera.org/sitemap~www~courses.xml'
    courses_body = requests.get(url=url)
    if courses_body.ok:
        logging.info('finish getting XML')
        return courses_body.content
    else:
        logging.error(courses_body.status_code)
        return None


def get_courses_url_list(courses_body, courses_num):
    logging.info(u'start parsing xml')
    courses_list = html.fromstring(courses_body)
    courses_url_list = courses_list.xpath('//loc/text()')
    logging.info(u'finish parsing xml')
    return courses_url_list[:courses_num]


def get_courses_html(courses_url_list):
    list_courses_html = []
    for course_url in courses_url_list:
        try:
            get_html = requests.get(course_url)
            html_page = get_html.text
            list_courses_html.append(html_page)
        except requests.exceptions:
            logging.error('{}: {} '.format(course_url, get_html.status_code))
    return list_courses_html


def get_course_info(list_courses_html):
    logging.info('start parsing')
    list_courses_json = []
    for html_page in list_courses_html:
        course_json = {}
        soup = BeautifulSoup(html_page, 'html.parser')
        for course_name in soup.find_all('h2',
                                         class_='headline-4-text '
                                                'course-title'):
            course_json['name'] = course_name.text
        for course_language in soup.find_all('div', class_='rc-Language'):
            course_json['language'] = course_language.text
        for course_start_date in soup.find_all('div',
                                               class_='startdate rc-StartDateString caption-text'):
            course_json['start_date'] = course_start_date.text
        for course_rating in soup.find_all('div',
                                           class_='ratings-text '
                                                  'bt3-visible-xs'):
            course_json['rating'] = course_rating.text
        course_json['duration'] = len(soup.find_all('div', class_='week'))
        list_courses_json.append(course_json)
        logging.info('Append: {}'.format(course_json['name']))
    logging.info('finish parsing')
    return list_courses_json


def output_courses_info_to_xlsx(list_courses_json, filename):
    logging.info('start generate .xlsx')
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Language', 'Rating', 'Start_date', 'Duration'])
    for row_number, course in enumerate(list_courses_json, start=2):
        try:
            ws.cell(row=row_number, column=1, value=course['name'])
            ws.cell(row=row_number, column=2, value=course['language'])
            ws.cell(row=row_number, column=3, value=course['rating'])
            ws.cell(row=row_number, column=4, value=course['start_date'])
            ws.cell(row=row_number, column=5, value=course['duration'])
        except KeyError:
            logging.error('{}: doesnt have some value'.format(course['name']))
    wb.save(filename + '.xlsx')
    logging.info('.xlsx done')


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('-a', '--amount', default=10, type=int)
    parser.add_argument('-n', '--name', default='devman', type=str)
    return parser.parse_args()


if __name__ == '__main__':
    args = parse_args()
    logging.info(u'coursera.py starts...')
    star_time = time.time()
    courses_content = get_courses_content()
    if courses_content:
        list_courses_url = get_courses_url_list(courses_content, args.amount)
        list_courses_html = get_courses_html(list_courses_url)
        list_courses_json = get_course_info(list_courses_html)
        output_courses_info_to_xlsx(list_courses_json, args.name)
    logging.info('Finish program: {}'.format(time.time() - star_time))

    print('\n \n Success! {}.xlsx is done. '.format(args.name))
